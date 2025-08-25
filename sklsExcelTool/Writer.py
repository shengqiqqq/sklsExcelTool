import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def writer_core(excel_path, df, sheet_name='Sheet1', startrow=0, startcol=0, index=True, header=True):
    """
    核心写入函数，将DataFrame写入Excel文件的指定工作表和单元格位置
    保留原有工作表内容，仅在指定位置写入新数据

    参数:
        excel_path (str): Excel文件路径
        df (pd.DataFrame): 要写入的DataFrame
        sheet_name (str): 工作表名称，默认为'Sheet1'
        startrow (int): 起始行索引（从0开始），默认为0
        startcol (int): 起始列索引（从0开始），默认为0
        index (bool): 是否写入索引，默认为True
        header (bool): 是否写入列名，默认为True
    """
    # 检查文件是否存在
    file_exists = os.path.exists(excel_path)

    if file_exists:
        # 加载现有工作簿
        wb = load_workbook(excel_path)

        # 如果工作表不存在则创建
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)

        # 获取工作表
        ws = wb[sheet_name]

        # 将DataFrame转换为行
        rows = dataframe_to_rows(df, index=index, header=header)

        # 写入数据到指定位置
        for r_idx, row in enumerate(rows, start=startrow + 1):  # openpyxl行索引从1开始
            for c_idx, value in enumerate(row, start=startcol + 1):  # openpyxl列索引从1开始
                ws.cell(row=r_idx, column=c_idx, value=value)

        # 保存工作簿
        wb.save(excel_path)
        wb.close()
    else:
        # 文件不存在时，使用pandas的to_excel创建新文件
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='w') as writer:
            df.to_excel(
                writer,
                sheet_name=sheet_name,
                startrow=startrow,
                startcol=startcol,
                index=index,
                header=header
            )

def write_to_excel(excel_path, df, sheet_name='Sheet1', startrow=0, startcol=0,
                   index=True, header=True, overwrite=False, clear_sheet=False):
    """
    外层函数，处理Excel文件的创建和写入，兼容已存在的文件

    参数:
        excel_path (str): Excel文件路径
        df (pd.DataFrame): 要写入的DataFrame
        sheet_name (str): 工作表名称，默认为'Sheet1'
        startrow (int): 起始行索引（从0开始），默认为0
        startcol (int): 起始列索引（从0开始），默认为0
        index (bool): 是否写入索引，默认为True
        header (bool): 是否写入列名，默认为True
        overwrite (bool): 如果文件存在，是否完全覆盖，默认为False
        clear_sheet (bool): 如果工作表存在，是否清空该工作表再写入，默认为False
    """
    # 确保目录存在
    dir_name = os.path.dirname(excel_path)
    if dir_name and not os.path.exists(dir_name):
        os.makedirs(dir_name, exist_ok=True)

    # 如果需要覆盖且文件存在，则先删除
    if overwrite and os.path.exists(excel_path):
        os.remove(excel_path)
        file_exists = False
    else:
        file_exists = os.path.exists(excel_path)

    # 如果需要清空工作表且文件存在
    if clear_sheet and file_exists:
        wb = load_workbook(excel_path)
        if sheet_name in wb.sheetnames:
            # 清空指定工作表
            ws = wb[sheet_name]
            ws.delete_rows(1, ws.max_row)
        wb.save(excel_path)
        wb.close()

    # 调用核心写入函数
    writer_core(excel_path, df, sheet_name, startrow, startcol, index, header)

    print(f"数据已成功写入到 {excel_path} 的 {sheet_name} 工作表，起始位置：行{startrow+1}，列{startcol+1}")

# 示例用法
if __name__ == "__main__":
    # 创建示例DataFrame
    data1 = {
        'Name': ['Alice', 'Bob', 'Charlie'],
        'Age': [25, 30, 35],
        'City': ['New York', 'London', 'Paris']
    }
    df1 = pd.DataFrame(data1)

    data2 = {
        'Product': ['Laptop', 'Phone', 'Tablet'],
        'Price': [999, 699, 299],
        'Stock': [10, 25, 15]
    }
    df2 = pd.DataFrame(data2)

    # 示例1: 写入新文件，从A1单元格开始
    write_to_excel('example_fixed.xlsx', df1, sheet_name='Data', index=False)

    # 示例2: 在同一工作表的C5单元格写入第二个DataFrame（不会清空原有内容）
    write_to_excel('example_fixed.xlsx', df2, sheet_name='Data',
                  startrow=4, startcol=2, header=True)

    # 示例3: 清空工作表后写入新数据
    write_to_excel('example_fixed.xlsx', df1, sheet_name='Data',
                  startrow=0, startcol=0, clear_sheet=True)