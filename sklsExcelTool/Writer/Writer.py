import os


def writer_when_opening(df, worksheet, startrow=1, startcol=1, index=True, header=True):
    '''
    打开时写入df

    :param df: 要写入的DataFrame
    :param worksheet: 需要Worksheet对象
    :param startrow: 开始行(从1开始)
    :param startcol: 开始列(从1开始)
    :param index: 是否写入索引
    :param header: 是否写入表头
    :return: 没有返回
    '''
    # 在函数内导入所需模块
    from openpyxl.utils.dataframe import dataframe_to_rows

    rows = dataframe_to_rows(df, index=index, header=header)

    for row_idx, row in enumerate(rows, start=startrow):
        for col_idx, value in enumerate(row, start=startcol):
            worksheet.cell(row=row_idx, column=col_idx, value=value)


def writer_to_excel(path, df, sheetname='Sheet1', startrow=1, startcol=1,
                      index=True, header=True, overwrite=False, clear_sheet=False, close=True):
    '''
    将DataFrame写入Excel文件

    :param path: Excel文件路径
    :param df: 要写入的DataFrame
    :param sheetname: 工作表名称
    :param startrow: 开始行(从1开始)
    :param startcol: 开始列(从1开始)
    :param index: 是否写入索引
    :param header: 是否写入表头
    :param overwrite: 是否覆盖现有文件
    :param clear_sheet: 是否清空工作表原有内容
    :param close: 是否关闭工作簿
    :return: 如果close为False，返回工作簿对象
    '''
    # 在函数内导入所需模块
    try:
        import openpyxl
        from openpyxl import load_workbook
    except ImportError as e:
        raise ImportError(f"请确保已安装openpyxl库: {e}")

    # 确保目录存在
    dir_name = os.path.dirname(path)
    if dir_name and not os.path.exists(dir_name):
        os.makedirs(dir_name, exist_ok=True)

    # 处理文件存在情况
    file_exists = os.path.exists(path)
    #print(file_exists)
    # 如果需要覆盖现有文件，则删除它
    if file_exists and overwrite:
        os.remove(path)
        file_exists = False

    # 获取工作簿
    if file_exists:
        workbook = load_workbook(path,read_only = False,)
    else:
        workbook = openpyxl.Workbook(write_only=False)#

    # 获取或创建工作表
    if sheetname in workbook.sheetnames:
        worksheet = workbook[sheetname]
        # 清空工作表内容
        if clear_sheet:
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.value = None
    else:
        # 如果是新工作簿且创建新工作表，删除默认的Sheet
        if not file_exists and 'Sheet' in workbook.sheetnames:
            del workbook['Sheet']
        worksheet = workbook.create_sheet(sheetname)

    # 执行写入
    writer_when_opening(
        df=df,
        worksheet=worksheet,
        startrow=startrow,
        startcol=startcol,
        index=index,
        header=header
    )

    # 保存并关闭工作簿
    if close:
        try:
            workbook.save(path)
        except Exception as e:
            print(f"保存文件时出错: {e}")
        finally:
            workbook.close()
    else:
        return workbook
